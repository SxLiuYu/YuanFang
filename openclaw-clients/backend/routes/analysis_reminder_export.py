from fastapi import APIRouter, Request
import logging, io
from datetime import datetime
from fastapi.responses import StreamingResponse
from services.analytics_service import (
    analytics_service, HealthScoreRequest, FinanceInsightRequest,
    AnomalyCheckRequest, ReportGenerateRequest
)
from services.anomaly_detection_service import AnomalyDetectionService
from services.trend_prediction_service import TrendPredictionService
from services.enhanced_reminder_service import get_enhanced_service
from services.personal_data_service import get_service as get_personal_service
from services.notification_parser_service import NotificationParserService
from services.natural_language_accounting_service import parse_accounting_text
from services.scene_automation_service import get_scene_service
from shared import success_response, error_response

router = APIRouter(tags=["analysis-reminder-export"])

_anomaly_service = None
_trend_service = None
_notification_parser = None

def get_anomaly_service():
    global _anomaly_service
    if _anomaly_service is None:
        _anomaly_service = AnomalyDetectionService()
    return _anomaly_service

def get_trend_service():
    global _trend_service
    if _trend_service is None:
        _trend_service = TrendPredictionService()
    return _trend_service

def get_notification_parser():
    global _notification_parser
    if _notification_parser is None:
        _notification_parser = NotificationParserService()
    return _notification_parser

# Analytics
@router.get("/api/v1/analytics/health/score")
async def get_health_score(user_id: str, device_id: str = None):
    try:
        result = await analytics_service.calculate_health_score(
            HealthScoreRequest(user_id=user_id, device_id=device_id))
        return success_response(result)
    except Exception as e:
        return error_response(str(e))

@router.get("/api/v1/analytics/finance/insights")
async def get_finance_insights(user_id: str, period_days: int = 30):
    try:
        result = await analytics_service.get_finance_insights(
            FinanceInsightRequest(user_id=user_id, period_days=period_days))
        return success_response(result)
    except Exception as e:
        return error_response(str(e))

@router.get("/api/v1/analytics/anomalies")
async def check_anomalies(user_id: str, data_type: str):
    try:
        result = await analytics_service.check_anomalies(
            AnomalyCheckRequest(user_id=user_id, data_type=data_type))
        return success_response(result)
    except Exception as e:
        return error_response(str(e))

@router.get("/api/v1/analytics/report/weekly")
async def generate_weekly_report(user_id: str):
    try:
        result = await analytics_service.generate_report(
            ReportGenerateRequest(user_id=user_id, report_type="weekly"))
        return success_response(result)
    except Exception as e:
        return error_response(str(e))

@router.get("/api/v1/analytics/report/monthly")
async def generate_monthly_report(user_id: str):
    try:
        result = await analytics_service.generate_report(
            ReportGenerateRequest(user_id=user_id, report_type="monthly"))
        return success_response(result)
    except Exception as e:
        return error_response(str(e))

# Anomaly Detection
@router.post("/api/v1/analysis/anomaly/health")
async def detect_health_anomaly(request: Request):
    body = await request.json()
    result = get_anomaly_service().detect_health_anomalies(
        profile_id=body.get('profile_id', 'default'), days=body.get('days', 7))
    return success_response(result)

@router.post("/api/v1/analysis/anomaly/expense")
async def detect_expense_anomaly(request: Request):
    body = await request.json()
    result = get_anomaly_service().detect_expense_anomalies(
        profile_id=body.get('profile_id', 'default'), days=body.get('days', 30))
    return success_response(result)

@router.post("/api/v1/analysis/anomaly/location")
async def detect_location_anomaly(request: Request):
    body = await request.json()
    result = get_anomaly_service().detect_location_anomalies(
        profile_id=body.get('profile_id', 'default'), days=body.get('days', 7))
    return success_response(result)

@router.post("/api/v1/analysis/anomaly/all")
async def detect_all_anomalies(request: Request):
    body = await request.json()
    result = get_anomaly_service().run_all_detections(
        profile_id=body.get('profile_id', 'default'))
    return success_response(result)

# Trend Prediction
@router.post("/api/v1/analysis/trend/health")
async def predict_health_trend(request: Request):
    body = await request.json()
    result = get_trend_service().predict_health_trends(
        profile_id=body.get('profile_id', 'default'), days=body.get('days', 30))
    return success_response(result)

@router.post("/api/v1/analysis/trend/expense")
async def predict_expense_trend(request: Request):
    body = await request.json()
    result = get_trend_service().predict_expense_trends(
        profile_id=body.get('profile_id', 'default'), days=body.get('days', 90))
    return success_response(result)

@router.post("/api/v1/analysis/trend/location")
async def predict_location_pattern(request: Request):
    body = await request.json()
    result = get_trend_service().predict_location_patterns(
        profile_id=body.get('profile_id', 'default'), days=body.get('days', 30))
    return success_response(result)

@router.post("/api/v1/analysis/trend/all")
async def predict_all_trends(request: Request):
    body = await request.json()
    result = get_trend_service().run_all_predictions(
        profile_id=body.get('profile_id', 'default'), days=body.get('days', 30))
    return success_response(result)

# Reminders
@router.post("/api/v1/reminder/create")
async def create_reminder(request: Request):
    body = await request.json()
    result = get_enhanced_service().create_reminder(
        title=body.get('title'), reminder_type=body.get('reminder_type', 'time'),
        trigger_time=body.get('trigger_time'), trigger_location=body.get('trigger_location'),
        description=body.get('description'), priority=body.get('priority', 'normal'),
        recurring=body.get('recurring'), profile_id=body.get('profile_id', 'default'))
    return success_response(result)

@router.post("/api/v1/reminder/medication")
async def create_medication_reminder(request: Request):
    body = await request.json()
    result = get_enhanced_service().create_medication_reminder(
        medication=body.get('medication'), times=body.get('times', ['08:00']),
        days=body.get('days'))
    return success_response(result)

@router.get("/api/v1/reminder/list")
async def list_reminders(profile_id: str = 'default', active_only: bool = True):
    return success_response(get_enhanced_service().get_reminders(profile_id, active_only))

@router.get("/api/v1/reminder/pending")
async def get_pending_reminders(profile_id: str = 'default'):
    return success_response(get_enhanced_service().get_pending_reminders(profile_id))

@router.get("/api/v1/reminder/upcoming")
async def get_upcoming_reminders(hours: int = 24, profile_id: str = 'default'):
    return success_response(get_enhanced_service().get_upcoming(hours, profile_id))

@router.post("/api/v1/reminder/{reminder_id}/trigger")
async def trigger_reminder(reminder_id: str):
    return success_response(get_enhanced_service().trigger(reminder_id))

@router.post("/api/v1/reminder/{reminder_id}/snooze")
async def snooze_reminder(reminder_id: str, minutes: int = 10):
    return success_response(get_enhanced_service().snooze(reminder_id, minutes))

@router.post("/api/v1/reminder/{reminder_id}/complete")
async def complete_reminder(reminder_id: str):
    return success_response(get_enhanced_service().complete(reminder_id))

@router.delete("/api/v1/reminder/{reminder_id}")
async def delete_reminder(reminder_id: str):
    return success_response(get_enhanced_service().delete(reminder_id))

@router.post("/api/v1/reminder/check-location")
async def check_location_reminders(request: Request):
    body = await request.json()
    result = get_enhanced_service().check_location_triggers(
        latitude=body.get('latitude'), longitude=body.get('longitude'),
        profile_id=body.get('profile_id', 'default'))
    return success_response(result)

@router.get("/api/v1/reminder/suggestions")
async def get_reminder_suggestions(profile_id: str = 'default'):
    return success_response(get_enhanced_service().get_suggestions(profile_id))

# Export
@router.get("/api/v1/export/health/pdf")
async def export_health_pdf(profile_id: str = 'default', days: int = 30):
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import cm
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    p.setFont("Helvetica-Bold", 24)
    p.drawString(2*cm, height - 2*cm, "Health Report")
    p.setFont("Helvetica", 12)
    p.drawString(2*cm, height - 3*cm, f"Profile: {profile_id}")
    p.drawString(2*cm, height - 3.5*cm, f"Period: Last {days} days")
    p.drawString(2*cm, height - 4*cm, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    personal_service = get_personal_service()
    summary = personal_service.get_health_summary()
    y = height - 6*cm
    p.setFont("Helvetica-Bold", 14)
    p.drawString(2*cm, y, "Today's Summary")
    y -= 1*cm
    p.setFont("Helvetica", 11)
    today = summary.get('today')
    if today:
        p.drawString(2*cm, y, f"Steps: {today.get('steps', 'N/A')}")
        y -= 0.5*cm
        p.drawString(2*cm, y, f"Heart Rate: {today.get('heart_rate', 'N/A')} bpm")
        y -= 0.5*cm
        p.drawString(2*cm, y, f"Sleep: {today.get('sleep_hours', 'N/A')} hours")
        y -= 0.5*cm
        p.drawString(2*cm, y, f"Calories: {today.get('calories', 'N/A')} kcal")
    else:
        p.drawString(2*cm, y, "No data for today")
    y -= 2*cm
    p.setFont("Helvetica-Bold", 14)
    p.drawString(2*cm, y, "Weekly Average")
    y -= 1*cm
    week = summary.get('week', {})
    p.setFont("Helvetica", 11)
    p.drawString(2*cm, y, f"Total Steps: {week.get('total_steps', 0)}")
    y -= 0.5*cm
    p.drawString(2*cm, y, f"Avg Heart Rate: {week.get('avg_heart_rate', 0)} bpm")
    y -= 0.5*cm
    p.drawString(2*cm, y, f"Avg Sleep: {week.get('avg_sleep', 0)} hours")
    y -= 0.5*cm
    p.drawString(2*cm, y, f"Total Calories: {week.get('total_calories', 0)} kcal")
    p.save()
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=health_report_{datetime.now().strftime('%Y%m%d')}.pdf"})

@router.get("/api/v1/export/finance/excel")
async def export_finance_excel(profile_id: str = 'default', month: str = None):
    import openpyxl
    from openpyxl.styles import Font, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Finance Report"
    ws['A1'] = "Finance Report"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f"Profile: {profile_id}"
    ws['A3'] = f"Month: {month or datetime.now().strftime('%Y-%m')}"
    ws['A4'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    personal_service = get_personal_service()
    summary = personal_service.get_payment_summary(month)
    ws['A6'] = "Summary"
    ws['A6'].font = Font(bold=True, size=12)
    ws['A7'] = "Total Expense"
    ws['B7'] = summary.get('total_expense', 0)
    ws['A8'] = "Total Income"
    ws['B8'] = summary.get('total_income', 0)
    ws['A9'] = "Transactions"
    ws['B9'] = summary.get('transaction_count', 0)
    ws['A11'] = "By Category"
    ws['A11'].font = Font(bold=True, size=12)
    ws['A12'] = "Category"
    ws['B12'] = "Amount"
    ws['A12'].font = Font(bold=True)
    ws['B12'].font = Font(bold=True)
    row = 13
    for category, amount in summary.get('by_category', {}).items():
        ws[f'A{row}'] = category
        ws[f'B{row}'] = amount
        row += 1
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=finance_report_{datetime.now().strftime('%Y%m%d')}.xlsx"})

@router.get("/api/v1/export/data/json")
async def export_data_json(profile_id: str = 'default'):
    personal_service = get_personal_service()
    data = {
        'profile_id': profile_id,
        'export_time': datetime.now().isoformat(),
        'health': {'summary': personal_service.get_health_summary(), 'history': personal_service.get_health_history(30)},
        'finance': {'summary': personal_service.get_payment_summary(), 'history': personal_service.get_payment_history(30)},
        'location': {'stats': personal_service.get_location_stats(7), 'history': personal_service.get_location_history(24)}
    }
    return success_response(data)

# Notification Parser
@router.post("/api/v1/notification/parse")
async def parse_notification(request: Request):
    body = await request.json()
    text = body.get('text', '')
    if not text:
        return error_response('请提供通知文本')
    parser = get_notification_parser()
    result = parser.parse(text)
    if result.get('type') == 'payment' and body.get('auto_record', False):
        payment = result.get('payment', {})
        personal_service = get_personal_service()
        personal_service.record_payment(
            amount=payment.get('amount', 0), merchant=payment.get('merchant', '未知'),
            platform=payment.get('platform'))
    return success_response(result)

@router.post("/api/v1/notification/verification-code")
async def extract_verification_code(request: Request):
    body = await request.json()
    text = body.get('text', '')
    parser = get_notification_parser()
    result = parser.parse(text)
    if result.get('type') == 'verification_code':
        return success_response({'code': result.get('verification_code', {}).get('code'),
                                  'service': result.get('verification_code', {}).get('service')})
    return success_response({'code': None})

# Natural Language Accounting
@router.post("/api/v1/accounting/parse")
async def parse_natural_language_accounting(request: Request):
    body = await request.json()
    text = body.get('text', '')
    if not text:
        return error_response('请输入记账内容')
    result = await parse_accounting_text(text)
    return success_response(result)

@router.post("/api/v1/accounting/quick-record")
async def quick_record_by_natural_language(request: Request):
    body = await request.json()
    text = body.get('text', '')
    if not text:
        return error_response('请输入记账内容')
    parsed = await parse_accounting_text(text)
    if not parsed.get('success'):
        return error_response('无法解析记账内容')
    data = parsed.get('data', {})
    personal_service = get_personal_service()
    result = personal_service.record_payment(
        amount=data.get('amount', 0), merchant=data.get('merchant'),
        category=data.get('category'), payment_type=data.get('type', 'expense'))
    return success_response({'parsed': data, 'recorded': result})

# Scene Automation
@router.post("/api/v1/scene/rules")
async def create_scene_rule(request: Request):
    body = await request.json()
    result = get_scene_service().create_rule(
        name=body.get('name'), scene_type=body.get('scene_type'),
        trigger_type=body.get('trigger_type'), trigger_config=body.get('trigger_config', {}),
        actions=body.get('actions', []))
    return success_response(result)

@router.get("/api/v1/scene/rules")
async def list_scene_rules(active_only: bool = True):
    return success_response(get_scene_service().get_rules(active_only))

@router.get("/api/v1/scene/rules/{rule_id}")
async def get_scene_rule(rule_id: str):
    rule = get_scene_service().get_rule(rule_id)
    if rule:
        return success_response(rule)
    return error_response('规则不存在')

@router.put("/api/v1/scene/rules/{rule_id}")
async def update_scene_rule(rule_id: str, request: Request):
    body = await request.json()
    return success_response(get_scene_service().update_rule(rule_id, **body))

@router.delete("/api/v1/scene/rules/{rule_id}")
async def delete_scene_rule(rule_id: str):
    return success_response(get_scene_service().delete_rule(rule_id))

@router.post("/api/v1/scene/rules/{rule_id}/activate")
async def activate_scene_rule(rule_id: str):
    return success_response(get_scene_service().activate_rule(rule_id))

@router.post("/api/v1/scene/rules/{rule_id}/deactivate")
async def deactivate_scene_rule(rule_id: str):
    return success_response(get_scene_service().deactivate_rule(rule_id))

@router.get("/api/v1/scene/templates")
async def get_scene_templates():
    return success_response(get_scene_service().get_templates())

@router.post("/api/v1/scene/templates/{template_name}")
async def create_from_template(template_name: str, request: Request):
    body = await request.json()
    return success_response(get_scene_service().create_from_template(template_name, body.get('config')))

@router.post("/api/v1/scene/trigger")
async def trigger_scene(request: Request):
    body = await request.json()
    scene_type = body.get('scene_type')
    if not scene_type:
        return error_response('请指定场景类型')
    return success_response(get_scene_service().trigger_scene(scene_type))

# Smart Home (alternate path)
@router.get("/api/v1/home/devices")
async def list_smart_devices():
    from services.smart_home_service import discover_devices
    result = await discover_devices()
    return success_response(result)

@router.post("/api/v1/home/devices/{device_id}/control")
async def control_smart_device(device_id: str, request: Request):
    from services.smart_home_service import control_device
    body = await request.json()
    action = body.get('action')
    value = body.get('value')
    if not action:
        return error_response('请指定操作')
    result = await control_device(device_id, action, value)
    return success_response(result)

@router.get("/api/v1/home/devices/{device_id}/status")
async def get_device_status_api(device_id: str):
    from services.smart_home_service import get_device_status
    result = await get_device_status(device_id)
    return success_response(result)